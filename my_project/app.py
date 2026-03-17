import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
import io
from datetime import datetime
from dateutil.relativedelta import relativedelta

# 페이지 설정
st.set_page_config(page_title="엑셀 데이터 연산 서비스", layout="wide")

st.title("엑셀 데이터 연산 및 다운로드 서비스")

# --- 강조된 업로드 안내 영역 ---
st.markdown('<p style="background-color: #FFFF00; color: black; font-weight: bold; padding: 12px; border-radius: 8px; font-size: 18px; border: 1px solid #CCAC00;">📂 관리파일_YYYYMMDD을 업로드하세요</p>', unsafe_allow_html=True)
uploaded_file = st.file_uploader("", type=["xlsx"])

if uploaded_file:
    with st.spinner('데이터를 분석하고 보고서를 생성하는 중입니다...'):
        try:
            # 1. 인풋 파일 읽기
            df_sheet1 = pd.read_excel(uploaded_file, sheet_name=0, header=None)
            df_sheet2 = pd.read_excel(uploaded_file, sheet_name=1, header=None)
            
            # --- [데이터 전처리 및 집계] ---
            # 설치연월 (H열: 인덱스 7)
            install_dates = pd.to_datetime(df_sheet1.iloc[:, 7], errors='coerce')
            
            # AA2 ~ AL2 데이터 (AA는 26번째 열, 인덱스 26 / 2행은 인덱스 1)
            row2_values = df_sheet1.iloc[1, 26:38].values 

            # 시트1 요약용 데이터 추출
            data_s1 = df_sheet1.iloc[:, [3, 4]].copy()
            data_s1.columns = ['status', 'item_name']
            data_s1['status'] = data_s1['status'].astype(str).str.strip()
            data_s1['item_name'] = data_s1['item_name'].astype(str).str.strip()
            
            s1_normal_counts = data_s1[data_s1['status'] == '정상']['item_name'].value_counts()
            s1_closed_counts = data_s1[data_s1['status'] == '폐업']['item_name'].value_counts()

            # 시트2 집계 데이터 추출
            data_s2 = df_sheet2.iloc[:, [3, 13, 14, 15, 17]].copy()
            data_s2.columns = ['item_d', 'item_n', 'sum_value', 'status', 'item_r']
            for col in ['item_d', 'item_n', 'status', 'item_r']:
                data_s2[col] = data_s2[col].astype(str).str.strip()
            data_s2['sum_value'] = pd.to_numeric(data_s2['sum_value'], errors='coerce').fillna(0)
            
            # 포함(Contains) 조건 설정
            is_out = data_s2['status'].str.contains('출고', na=False)
            is_hold = data_s2['status'].str.contains('보유', na=False)

            # 집계 사전 생성
            s2_d_total = data_s2['item_d'].value_counts()
            s2_d_out = data_s2[is_out]['item_d'].value_counts()
            s2_d_hold = data_s2[is_hold]['item_d'].value_counts()
            s2_d_sum = data_s2.groupby('item_d')['sum_value'].sum()
            s2_n_out = data_s2[is_out]['item_n'].value_counts()
            s2_n_hold = data_s2[is_hold]['item_n'].value_counts()
            s2_r_counts = data_s2['item_r'].value_counts()

            # 2. 템플릿 처리
            base_path = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(base_path, "templates", "template_B.xlsx")
            
            if os.path.exists(template_path):
                wb = load_workbook(template_path)
                ws = wb.active
                
                # --- 상단 요약 (7행) ---
                ws['B7'] = df_sheet1.iloc[5:10000, 3].count()
                ws['D7'] = (df_sheet1.iloc[5:10000, 3].astype(str).str.strip() == '정상').sum()
                ws['F7'] = (df_sheet1.iloc[5:10000, 3].astype(str).str.strip() == '폐업').sum()
                
                s2_status_7 = df_sheet2.iloc[4:10000, 15].astype(str)
                ws['I7'] = df_sheet2.iloc[4:10000, 15].count()
                ws['K7'] = s2_status_7.str.contains('출고').sum()
                ws['M7'] = s2_status_7.str.contains('보유').sum()
                
                # --- 상세 기입 (11~16행 / 21~23행) ---
                for row_num in range(11, 17):
                    if row_num <= 15:
                        key_b = str(ws[f'B{row_num}'].value).strip() if ws[f'B{row_num}'].value else ""
                        if key_b:
                            ws[f'D{row_num}'] = s1_normal_counts.get(key_b, 0) + s1_closed_counts.get(key_b, 0)
                            ws[f'E{row_num}'] = s1_normal_counts.get(key_b, 0)
                            ws[f'F{row_num}'] = s1_closed_counts.get(key_b, 0)
                    
                    key_i = str(ws[f'I{row_num}'].value).strip() if ws[f'I{row_num}'].value else ""
                    if key_i:
                        ws[f'J{row_num}'] = s2_d_total.get(key_i, 0)
                        ws[f'K{row_num}'] = s2_d_out.get(key_i, 0)
                        ws[f'L{row_num}'] = s2_d_hold.get(key_i, 0)
                        ws[f'M{row_num}'] = s2_d_sum.get(key_i, 0)

                for row_num in range(21, 24):
                    key_b_21 = str(ws[f'B{row_num}'].value).strip() if ws[f'B{row_num}'].value else ""
                    if key_b_21:
                        ws[f'D{row_num}'] = s2_n_out.get(key_b_21, 0)
                        ws[f'F{row_num}'] = s2_n_hold.get(key_b_21, 0)
                    if row_num < 23:
                        key_i_21 = str(ws[f'I{row_num}'].value).strip() if ws[f'I{row_num}'].value else ""
                        if key_i_21:
                            ws[f'K{row_num}'] = s2_r_counts.get(key_i_21, 0)

                # --- [28행 데이터 기입 (AA2~AL2)] ---
                for i, val in enumerate(row2_values):
                    ws.cell(row=28, column=3 + i).value = val

                # --- [30행 설치연월 집계] ---
                start_date = datetime(2025, 3, 1)
                for i in range(12):
                    target_month = start_date + relativedelta(months=i)
                    count = ((install_dates.dt.year == target_month.year) & 
                             (install_dates.dt.month == target_month.month)).sum()
                    ws.cell(row=30, column=3 + i).value = count
                
                # 수식 계산 엔진 활성화 (스타일 보존에 도움)
                wb.calculation.calcMode = 'auto'
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                download_name = f"{os.path.splitext(uploaded_file.name)[0]}_Report.xlsx"
                st.success("✅ 보고서 생성이 완료되었습니다!")
                st.download_button("📊 결과 파일 다운로드", output, download_name)
            else:
                st.error("❌ 템플릿 파일(template_B.xlsx)을 찾을 수 없습니다.")
        except Exception as e:
            st.error(f"⚠️ 오류가 발생했습니다: {e}")
